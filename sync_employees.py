#!/usr/bin/env python3
# sync_employees.py - Синхронизация сотрудников из Excel

import sqlite3
import openpyxl
import os
import re
from datetime import datetime
from config import EMPLOYEES_DB_PATH

DB_PATH = EMPLOYEES_DB_PATH
EXCEL_PATH = os.getenv('EXCEL_PATH', '/data/phonebook.xlsx')

def clean_name(name):
    if not name:
        return None
    name = str(name).replace('\n', ' ').strip()
    name = re.sub(r'\s+', ' ', name)
    return name.upper()

def is_valid_email(email):
    if not email:
        return False
    email = str(email).strip()
    return re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email)

def sync_from_excel():
    print(f"🔄 Синхронизация начата: {datetime.now()}")

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Файл {EXCEL_PATH} не найден")
        return False

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        sheet = wb.active
        print(f"📖 Прочитан лист: {sheet.title}")
    except Exception as e:
        print(f"❌ Ошибка чтения Excel: {e}")
        return False

    # Ищем заголовки
    headers = {}
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value:
            headers[cell_value] = col

    if 'ФИО' not in headers or 'E-mail' not in headers:
        print("❌ Не найдены колонки 'ФИО' или 'E-mail'")
        print(f"📋 Найдены колонки: {list(headers.keys())}")
        return False

    # Собираем сотрудников (email — уникальный ключ)
    employees = {}
    for row in range(2, sheet.max_row + 1):
        name = clean_name(sheet.cell(row=row, column=headers['ФИО']).value)
        email = sheet.cell(row=row, column=headers['E-mail']).value

        if not name or not is_valid_email(email):
            continue

        employees[email] = name

    print(f"👤 Найдено {len(employees)} уникальных сотрудников")

    # Подключаемся к БД
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Создаём таблицу, если её нет
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL
        )
    ''')
    
    # Добавляем индексы для быстрого поиска
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_full_name ON employees(full_name)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_email ON employees(email)')

    # Очищаем таблицу
    cursor.execute("DELETE FROM employees")
    print("🗑️ Старые данные удалены")

    # Заполняем данными
    added = 0
    for email, name in employees.items():
        cursor.execute('''
            INSERT INTO employees (full_name, email)
            VALUES (?, ?)
        ''', (name, email))
        added += 1

    conn.commit()
    conn.close()

    print(f"✅ Синхронизация завершена! Добавлено {added} сотрудников")

    # Проверка
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM employees")
    count = cursor.fetchone()[0]
    conn.close()
    print(f"📊 В БД {count} записей")

    return True

if __name__ == '__main__':
    sync_from_excel()
